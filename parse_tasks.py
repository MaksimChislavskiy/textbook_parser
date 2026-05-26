import os
import re
import openpyxl
from docx import Document

INPUT_FOLDER = 'input_data'
OUTPUT_FOLDER = 'output'
TARGET_FILE = 'tekstovye_zadachi_po_matematike.docx'
TOC_FILE = 'результат_tekstovye_zadachi_po_matematike.xlsx'  # файл с оглавлением

# Задача с буквой: "1. а)", "3. 1)", "207*. 2)"
TASK_WITH_LETTER_RE = re.compile(r'^(\d+\*?)\.\s*\S?\s*([а-я]\)|\d\))\s*')

# Продолжение задачи: "б)", "2)" (без номера задачи)
TASK_CONTINUE_RE = re.compile(r'^\s*\S?\s*([а-я]\)|\d\))\s*')

# Задача без буквы: "6. Первая бригада..."
TASK_NO_LETTER_RE = re.compile(r'^(\d+\*?)\.\s+\S?\s*([А-Я].+)')

# Раздел: 1-3 слова, без номера страницы
SECTION_IN_TASKS_RE = re.compile(
    r'^(\d+)\.\s+([А-Я][а-я]+(?:\s+[а-я]+){0,2})\s*$'
)

# Подраздел: цифра.цифра, любой текст
SUBSECTION_RE = re.compile(
    r'^(\d+\.\d+)\.?\s+(.+?)\s*$'
)

# Варианты через точку с запятой: "а) 1 рубля; б) 1 метра"
INLINE_VARIANTS_RE = re.compile(r';\s*[а-я\d]\)')

# Два и более подпункта в одной строке: "1) ... 2) ..."
MULTI_VARIANTS_RE = re.compile(r'[а-я\d]\).*[а-я\d]\)')


def load_toc_mapping():
    """Загружает оглавление и создаёт словарь: название_подраздела → id"""
    toc_path = os.path.join(OUTPUT_FOLDER, TOC_FILE)
    if not os.path.exists(toc_path):
        print(f"Предупреждение: файл {toc_path} не найден. Параграфы не будут заполнены.")
        return {}

    wb = openpyxl.load_workbook(toc_path)
    ws = wb.active

    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        toc_id = row[0]
        name = row[1]
        parent = row[2]

        if parent != 0 and name:
            clean_name = name.strip().rstrip('.')
            mapping[clean_name] = toc_id

    # print(f"Загружено {len(mapping)} подразделов из оглавления")
    return mapping


def get_paragraph_id(section_name, toc_mapping):
    """Возвращает id параграфа по названию подраздела"""
    clean_name = section_name.strip().rstrip('.')
    return toc_mapping.get(clean_name, 0)


if TARGET_FILE:
    doc_path = os.path.join(INPUT_FOLDER, TARGET_FILE)
    if not os.path.exists(doc_path):
        raise FileNotFoundError(f'Файл {doc_path} не найден')
else:
    raise ValueError('Укажите TARGET_FILE')

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Загружаем соответствие названий параграфов их id
toc_mapping = load_toc_mapping()

doc = Document(doc_path)
all_lines = []
for paragraph in doc.paragraphs:
    line = paragraph.text.strip()
    if line:
        all_lines.append(line)

tasks = []
current_task = None
current_task_num = None
current_paragraph_id = 0
mode = 'toc'

# print("Начало парсинга...")

for line in all_lines:
    # === Смена режимов ===
    if line == 'Оглавление' and mode != 'answers':
        mode = 'toc'
        continue

    if line.startswith('Ответы и советы'):
        mode = 'answers'
        continue

    if line == 'Оглавление' and mode == 'answers':
        break

    # === Парсинг содержания ===
    if mode == 'toc':
        subsection_match = SUBSECTION_RE.match(line)
        if subsection_match:
            section_name = subsection_match.group(2)
            new_id = get_paragraph_id(section_name, toc_mapping)
            if new_id:
                current_paragraph_id = new_id
                # print(f"  → Параграф {current_paragraph_id}: {section_name}")

        if SECTION_IN_TASKS_RE.match(line) or SUBSECTION_RE.match(line):
            continue
        else:
            mode = 'tasks'
            # print("Переход в режим парсинга задач")

    # === Парсинг задач ===
    if mode == 'tasks':
        if SECTION_IN_TASKS_RE.match(line):
            continue

        subsection_match = SUBSECTION_RE.match(line)
        if subsection_match:
            section_name = subsection_match.group(2)
            new_id = get_paragraph_id(section_name, toc_mapping)
            if new_id:
                current_paragraph_id = new_id
                # print(f"  → Обновлён параграф {current_paragraph_id}: {section_name}")
            continue

        full_match = TASK_WITH_LETTER_RE.match(line)
        continue_match = TASK_CONTINUE_RE.match(line)
        no_letter_match = TASK_NO_LETTER_RE.match(line)

        if full_match:
            text_after = line[full_match.end():].strip()

            if INLINE_VARIANTS_RE.search(text_after) or MULTI_VARIANTS_RE.search(text_after):
                if current_task:
                    current_task['task'] = current_task['task'] + " " + line
                continue

            if current_task and not current_task['task'].rstrip().endswith(':'):
                current_task['task'] = current_task['task'].strip()
                tasks.append(current_task)
            elif current_task:
                current_task['task'] = current_task['task'] + " " + line
                current_task_num = full_match.group(1)
                current_task['id_tasks_book'] = f"{current_task_num}.{full_match.group(2)}"
                continue

            current_task_num = full_match.group(1)
            sub_letter = full_match.group(2)
            current_task = {
                'id_tasks_book': f"{current_task_num}.{sub_letter}",
                'task': text_after,
                'classes': '5,6',
                'level': '1',
                'paragraph_id': current_paragraph_id,
                'topic_id': 1,
            }

        elif continue_match and current_task_num:
            text_after = line[continue_match.end():].strip()

            if INLINE_VARIANTS_RE.search(text_after) or MULTI_VARIANTS_RE.search(text_after):
                if current_task:
                    current_task['task'] = current_task['task'] + " " + line
                continue

            if current_task and not current_task['task'].rstrip().endswith(':'):
                current_task['task'] = current_task['task'].strip()
                tasks.append(current_task)
            elif current_task:
                current_task['task'] = current_task['task'] + " " + line
                continue

            sub_letter = continue_match.group(1)
            current_task = {
                'id_tasks_book': f"{current_task_num}.{sub_letter}",
                'task': text_after,
                'classes': '5,6',
                'level': '1',
                'paragraph_id': current_paragraph_id,
                'topic_id': 1,
            }

        elif no_letter_match:
            if current_task and not current_task['task'].rstrip().endswith(':'):
                current_task['task'] = current_task['task'].strip()
                tasks.append(current_task)
            elif current_task:
                current_task['task'] = current_task['task'] + " " + line
                continue

            current_task_num = no_letter_match.group(1)
            text = no_letter_match.group(2).strip()
            current_task = {
                'id_tasks_book': f'{current_task_num})',
                'task': text,
                'classes': '5,6',
                'level': '1',
                'paragraph_id': current_paragraph_id,
                'topic_id': 1,
            }

        else:
            if current_task:
                current_task['task'] = current_task['task'] + " " + line

# Сохраняем последнюю задачу
if current_task:
    current_task['task'] = current_task['task'].strip()
    tasks.append(current_task)

# === Excel ===
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "tasks"
ws.append(["id_tasks_book", "task", "answer", "classes", "topic_id", "level", "paragraph_id"])

for task in tasks:
    ws.append([
        task.get('id_tasks_book', ''),
        task.get('task', ''),
        task.get('answer', ''),
        task.get('classes', '5,6'),
        task.get('topic_id', 1),
        task.get('level', '1'),
        task.get('paragraph_id', 0),
    ])

output_name = f"задачи_{TARGET_FILE.replace('.docx', '.xlsx')}"
output_path = os.path.join(OUTPUT_FOLDER, output_name)

try:
    wb.save(output_path)
    print(f"\n Готово! {len(tasks)} задач сохранено в {output_path}")

    # Статистика
    # para_counts = {}
    # for task in tasks:
    #     para_id = task.get('paragraph_id', 0)
    #     para_counts[para_id] = para_counts.get(para_id, 0) + 1

    # print("\n📊 Распределение задач по параграфам:")
    # for para_id, count in sorted(para_counts.items()):
        # if para_id == 0:
        #     print(f"  ⚠️ Без параграфа: {count} задач")
        # else:
        #     print(f"  Параграф {para_id}: {count} задач")
except PermissionError:
    print(f"Ошибка: закройте файл {output_path} перед повторным запуском")