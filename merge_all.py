import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

INPUT_FOLDER = 'output'
TARGET_FILE = 'tekstovye_zadachi_po_matematike.docx'

TASKS_FILE = os.path.join(INPUT_FOLDER, f'задачи_{TARGET_FILE.replace(".docx", ".xlsx")}')
ANSWERS_FILE = os.path.join(INPUT_FOLDER, f'ответы_{TARGET_FILE.replace(".docx", ".xlsx")}')
TOC_FILE = os.path.join(INPUT_FOLDER, f'результат_{TARGET_FILE.replace(".docx", ".xlsx")}')
OUTPUT_FILE = os.path.join(INPUT_FOLDER, '5_6_class_shevkin_merged.xlsx')


def merge_all():
    # print("Загрузка данных...")
    tasks_df = pd.read_excel(TASKS_FILE, sheet_name='tasks')
    answers_df = pd.read_excel(ANSWERS_FILE, sheet_name='answers')
    toc_df = pd.read_excel(TOC_FILE, sheet_name='table_of_contents')

    print(f"Загружено: {len(tasks_df)} задач, {len(answers_df)} ответов, {len(toc_df)} разделов")

    # Объединение задач с ответами
    merged_df = tasks_df.merge(answers_df, on='id_tasks_book', how='left')

    if 'answer_x' in merged_df.columns and 'answer_y' in merged_df.columns:
        merged_df['answer'] = merged_df['answer_y'].fillna(merged_df['answer_x'])
        merged_df = merged_df.drop(columns=['answer_x', 'answer_y'])
    elif 'answer' not in merged_df.columns:
        merged_df['answer'] = ''

    # Порядок колонок
    column_order = ['id_tasks_book', 'task', 'answer', 'classes', 'paragraph_id', 'topic_id', 'level']
    for col in column_order:
        if col not in merged_df.columns:
            merged_df[col] = ''
    merged_df = merged_df[column_order]

    # Авторский лист
    author_data = {
        'name': ['Текстовые задачи по математике. 5–6 классы / А. В. Шевкин. — 3-е изд., перераб. — М. : Илекса, 2024. — 160 с. : ил.'],
        'author': ['А. В. Шевкин'],
        'description': ['Сборник включает текстовые задачи по разделам школьной математики: натуральные числа, дроби, пропорции, проценты, уравнения. Ко многим задачам даны ответы или советы с чего начать решения.'],
        'topic_id': [1],
        'classes': ['5;6']
    }
    author_df = pd.DataFrame(author_data)

    # Сохранение
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        merged_df.to_excel(writer, sheet_name='tasks', index=False)
        author_df.to_excel(writer, sheet_name='author', index=False)
        toc_df.to_excel(writer, sheet_name='table_of_contents', index=False)

    # Форматирование
    wb = load_workbook(OUTPUT_FILE)
    for sheet in ['tasks', 'author', 'table_of_contents']:
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            for cell in ws[1]:
                cell.font = Font(bold=True)
    wb.save(OUTPUT_FILE)

    print(f"\n Готово! Файл сохранён: {OUTPUT_FILE}")
    print(f"   - tasks: {len(merged_df)} задач")
    print(f"   - table_of_contents: {len(toc_df)} разделов")
    print(f"   - author: 1 запись")


if __name__ == '__main__':
    merge_all()